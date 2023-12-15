import textacy
from textacy import text_stats as ts
import pandas as pd
from openpyxl import load_workbook
# remember to do python -m spacy download en_core_web_sm

sheet = "spreadsheet.xlsx"

def getGradeLevel(text):
    doc = textacy.make_spacy_doc(text, lang="en_core_web_sm")
    return ts.readability.flesch_kincaid_grade_level(doc)

df = pd.read_excel(sheet, engine='openpyxl', dtype=object, header=None)
cellTexts = df.values.tolist()
cellTexts = cellTexts[1:]

grades = []
for cellText in cellTexts:
    grades.append(getGradeLevel(cellText[0]))

wb = load_workbook(filename=sheet)
ws = wb['Sheet1']

for i in range(len(grades)):
    ws.cell(row=i+2, column=2).value = grades[i]

wb.save(filename=sheet)
